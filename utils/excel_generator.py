# -*- coding: utf-8 -*-
"""
Excel generator for Trademark Search results
Maintains EXACT same formatting and image embedding as desktop version
"""

import os
import time
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage

class ExcelGenerator:
    def __init__(self):
        pass
    
    def generate_excel(self, search_results):
        """Generate Excel file with embedded images - EXACT same format as desktop version"""
        if not search_results:
            raise Exception("No search results to export")
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trademark Search Results"
        
        # EXACT same styles as desktop version
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # EXACT same headers as desktop version
        headers = [
            'S.No', 'Application Number', 'Wordmark', 'Proprietor', 
            'Class', 'Status', 'Image', 'Search Parameters'
        ]
        
        # Add headers - SAME formatting as desktop
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border
        
        # EXACT same column widths as desktop version
        column_widths = [8, 20, 25, 30, 8, 15, 25, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Process results - SAME logic as desktop version
        for idx, result in enumerate(search_results, 1):
            row = idx + 1
            
            # SAME row height for images - 60 points as in desktop version
            ws.row_dimensions[row].height = 60
            
            # Add data - EXACT same fields as desktop version
            data = [
                idx,
                result.get('Application_Number', ''),
                result.get('Wordmark', ''),
                result.get('Proprietor', ''),
                result.get('Class', ''),
                result.get('Status', ''),
                '',  # Image column will be filled with actual image
                self._format_search_params(result)
            ]
            
            # Add text data
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
        
        # Add images - EXACT same method as desktop version
        self._embed_images(ws, search_results)
        
        # Add summary section - SAME as desktop version
        self._add_summary_section(ws, search_results)
        
        # SAME auto-filter as desktop version
        ws.auto_filter.ref = f"A1:H{len(search_results) + 1}"
        
        # Save to BytesIO for web download - equivalent to desktop file save
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _embed_images(self, ws, search_results):
        """Embed images in Excel cells - EXACT same method as desktop version"""
        for idx, result in enumerate(search_results, 1):
            row = idx + 1
            image_data = result.get('Image_Data')
            
            if image_data:
                try:
                    # Create PIL Image from bytes - SAME as desktop
                    pil_image = PILImage.open(BytesIO(image_data))
                    
                    # EXACT same image processing as desktop version
                    # Resize to fit cell (approximately 100x50 pixels)
                    max_width, max_height = 100, 50
                    pil_image.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
                    
                    # Save processed image to BytesIO - SAME method
                    img_buffer = BytesIO()
                    
                    # Convert to RGB if necessary - SAME logic as desktop
                    if pil_image.mode in ('RGBA', 'P'):
                        pil_image = pil_image.convert('RGB')
                    
                    pil_image.save(img_buffer, format='JPEG', quality=85)
                    img_buffer.seek(0)
                    
                    # Create Excel image - SAME as desktop version
                    excel_img = ExcelImage(img_buffer)
                    
                    # SAME positioning logic as desktop
                    excel_img.anchor = f'G{row}'  # Column G (Image column)
                    
                    # Add to worksheet - SAME method
                    ws.add_image(excel_img)
                    
                except Exception as e:
                    print(f"Error processing image for row {row}: {str(e)}")
                    # Add error text in image cell - SAME fallback as desktop
                    ws.cell(row=row, column=7, value="Image Error")
    
    def _format_search_params(self, result):
        """Format search parameters - EXACT same format as desktop version"""
        params = []
        
        # SAME parameter formatting as desktop
        if result.get('Search_Wordmark'):
            params.append(f"Wordmark: {result['Search_Wordmark']}")
        
        if result.get('Search_Class'):
            params.append(f"Class: {result['Search_Class']}")
        
        if result.get('Search_Filter'):
            params.append(f"Filter: {result['Search_Filter']}")
        
        if result.get('Search_Date'):
            params.append(f"Date: {result['Search_Date']}")
        
        return "\n".join(params)
    
    def _add_summary_section(self, ws, search_results):
        """Add summary section - EXACT same as desktop version"""
        # Find the last data row
        last_row = len(search_results) + 1
        
        # Add some space
        summary_start_row = last_row + 3
        
        # SAME summary formatting as desktop
        summary_font = Font(name='Arial', size=11, bold=True, color='366092')
        summary_data_font = Font(name='Arial', size=10)
        
        # Summary headers - EXACT same as desktop
        summary_data = [
            ('Search Summary', ''),
            ('Total Results Found:', len(search_results)),
            ('Search Date:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ('Results with Images:', sum(1 for r in search_results if r.get('Image_Data'))),
            ('Results without Images:', sum(1 for r in search_results if not r.get('Image_Data')))
        ]
        
        # Add summary data - SAME formatting as desktop
        for i, (label, value) in enumerate(summary_data):
            row = summary_start_row + i
            
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = summary_font
            
            if value != '':
                value_cell = ws.cell(row=row, column=2, value=value)
                value_cell.font = summary_data_font
        
        # SAME search parameters section as desktop
        if search_results:
            first_result = search_results[0]
            params_start_row = summary_start_row + len(summary_data) + 2
            
            ws.cell(row=params_start_row, column=1, value="Search Parameters Used:").font = summary_font
            
            param_details = [
                f"Wordmark: {first_result.get('Search_Wordmark', 'N/A')}",
                f"Class: {first_result.get('Search_Class', 'N/A')}",
                f"Filter Type: {first_result.get('Search_Filter', 'N/A')}"
            ]
            
            for i, param in enumerate(param_details):
                ws.cell(row=params_start_row + i + 1, column=1, value=param).font = summary_data_font